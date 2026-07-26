#!/usr/bin/env python3
"""
지하철 계획 시각표 파서 — 정시성 판정의 대조군 (docs §3.3.1 · §8.1 · §9 #1·#2)

실측(subway-*.jsonl 의 recptnDt·통과시각)과 대조할 **계획 시각표**를 정규화한다.
정시성(§8 #1)은 계획 vs 실측이므로 계획 쪽이 없으면 판정이 안 된다.

소스가 둘이고 형식이 다르다:
  ① 신분당선 — 운영사(네오트랜스) 공식 PDF (§3.3.1)
     `역별 열차시각표_평일.pdf` · `_공휴일.pdf` — 16역 × 양방향. 열차번호가 **없다**
     (시·분 배열뿐) → 판정은 시각 근접 매칭으로만 가능(§8.1 ①).
     ⚠️ 신분당선은 이 시각표가 **선행 조건**이다: recptnDt 가 87% 미래(예측성,
     §8.1 ⑤ 가)라, 이 계획표와 대조해 "recptnDt 가 계획을 복사한 것인지"를
     먼저 가려야 정시성 판정에 쓸 수 있다. 계획과 recptnDt 가 거의 일치하면
     그 값은 실측이 아니라 시각표 기반 예측이고, 대조는 순환이 된다.
  ② 1~9호선 — 서울교통공사 파일 15098251 (§3.3.1, 별도 CLI: `stcs`)
     32MB·cp949·405역·DAY/SAT/END. 열차코드(S902/K802)가 있으나 실측 btrainNo
     (숫자)와 형식이 어긋나 조인 불가(§8.1 ①) → 역시 시각 근접 매칭.
     ⚠️ CSV 는 저장소에 없다. data.go.kr 15098251 fileData 에서 받아
     (로그인 불요) 경로를 넘길 것.

산출: 정규화 레코드 리스트. 한 레코드 = 한 역의 한 방면 한 출발(또는 통과) 시각.
  {"line","station","daytype","direction","updn","h","m","dest"}
  - daytype: 평일 / 토요일 / 공휴일 (소스 요일종을 그대로 보존)
  - direction: 방면 라벨(예 "광교방면"). updn: 상행/하행 정규화(0 상행 / 1 하행)
  - dest: 단축운행 행선지(예 "정자"), 전 구간 운행이면 None

사용:
  python3 subway_timetable.py sinbundang "역별 열차시각표_평일.pdf" --daytype 평일
  python3 subway_timetable.py sinbundang <_평일.pdf> <_공휴일.pdf> --json out.json
  python3 subway_timetable.py stcs 서울교통공사_열차시각표.csv          # 15098251
"""

import json
import re
import subprocess
import sys


# 신분당선 방면 → 상하행. 신사방면 = 상행(서울 방향), 광교방면 = 하행.
SBL_UPDN = {"신사방면": 0, "광교방면": 1}


def _pdftotext(path):
    """PDF → -layout 텍스트. pdftotext(poppler)가 없으면 명확히 안내하고 중단."""
    try:
        r = subprocess.run(["pdftotext", "-layout", path, "-"],
                           capture_output=True, text=True, timeout=60)
    except FileNotFoundError:
        sys.exit("pdftotext 가 없다 — poppler 설치 필요 (macOS: brew install poppler / "
                 "ubuntu: apt install poppler-utils). §9 #1 은 `pdftotext -layout` 전제다.")
    if r.returncode != 0:
        sys.exit(f"pdftotext 실패: {r.stderr.strip()[:200]}")
    return r.stdout


def _minute_token(tok):
    """'54(정자)' → (54, '정자'), '48' → (48, None). 분이 아니면 None."""
    m = re.match(r"^(\d{1,2})(?:\(([^)]+)\))?$", tok)
    if not m:
        return None
    return int(m.group(1)), m.group(2)


def parse_sinbundang(pdf_path, daytype):
    """신분당선 PDF 한 개(평일 또는 공휴일) → 정규화 레코드 리스트.

    페이지마다 한 역이고, `시`(hour) 컬럼을 기준으로 좌=신사방면·우=광교방면이다
    (종착역은 한 방면만 있다 — 신사=광교방면만, 광교=신사방면만). 컬럼 위치로
    가르므로 방면 라벨이 그 줄에 없어도(데이터 줄) 정확히 분리된다.
    """
    txt = _pdftotext(pdf_path)
    # 폼피드(\f) 또는 "Page No" 로 페이지를 가른다. -layout 은 페이지 사이에 \f 를 남긴다.
    pages = txt.split("\f")
    out = []
    for page in pages:
        lines = page.split("\n")
        station = None
        hour_col = None       # 헤더의 '시' 문자 컬럼 — 데이터 줄의 hour 기준
        dirs = {}             # 컬럼 기준: 'left'/'right' → 방면 라벨
        for i, line in enumerate(lines):
            mstat = re.search(r"(\S+)\s+열차시각표", line)
            if mstat:
                station = mstat.group(1)
                continue
            # 컬럼 헤더 줄 — 단독 '시' 가 있고 방면 라벨이 그 좌/우에 온다
            if hour_col is None and re.search(r"(^|\s)시(\s|$)", line) and "방면" in line:
                hcol = line.index("시", )
                # '시' 가 방면 라벨의 일부가 아닌, 컬럼 헤더의 시인지 — 방면 위치로 판별
                hour_col = _standalone_si_col(line)
                for mm in re.finditer(r"\S*방면", line):
                    (dirs.__setitem__("left", mm.group()) if mm.start() < hour_col
                     else dirs.__setitem__("right", mm.group()))
                continue
            if station is None or hour_col is None:
                continue
            # 데이터 줄 — hour_col 에 가장 가까운 토큰이 hour, 좌/우가 분·방면
            toks = [(m.start(), m.group()) for m in re.finditer(r"\S+", line)]
            if not toks:
                continue
            hj = min(range(len(toks)), key=lambda j: abs(toks[j][0] - hour_col))
            hcol, htok = toks[hj]
            if abs(hcol - hour_col) > 4 or not re.fullmatch(r"\d{1,2}", htok):
                continue                          # hour 컬럼에 정렬된 2자리 정수가 아니다
            hour = int(htok)
            for col, tok in toks:
                if tok is htok and col == hcol:
                    continue
                mt = _minute_token(tok)
                if mt is None:
                    continue
                minute, dest = mt
                side = "left" if col < hour_col else "right"
                direction = dirs.get(side)
                if not direction:
                    continue                      # 종착역의 빈 방면 쪽
                out.append({
                    "line": "신분당선", "station": station, "daytype": daytype,
                    "direction": direction, "updn": SBL_UPDN.get(direction),
                    "h": hour, "m": minute, "dest": dest,
                })
    return out


def _standalone_si_col(line):
    """방면 라벨('...방면')에 속하지 않은 단독 '시' 의 컬럼. 컬럼 헤더 판별용."""
    for m in re.finditer("시", line):
        c = m.start()
        # 앞뒤가 공백(또는 끝)이면 단독 '시' (방면의 '시' 아님 — 방면엔 '시' 없음)
        before = line[c-1] if c > 0 else " "
        after = line[c+1] if c+1 < len(line) else " "
        if before == " " and after == " ":
            return c
    return line.index("시")


def summarize(records):
    """역·방면별 열차 수와 첫차/막차 — 파싱이 온전한지 눈으로 확인."""
    by = {}
    for r in records:
        by.setdefault((r["station"], r["direction"]), []).append((r["h"], r["m"]))
    print(f"레코드 {len(records):,}건 · 역×방면 {len(by)}개")
    dests = {r["dest"] for r in records if r["dest"]}
    if dests:
        print(f"단축운행 행선지: {', '.join(sorted(dests))}")
    print(f"\n{'역':<12}{'방면':<10}{'열차':>5}  {'첫차':>6} {'막차':>6}")
    for (st, d), ts in by.items():
        ts.sort()
        # 막차가 자정 넘김(00~03시)이면 24+ 로 봐서 정렬상 뒤로
        def key(t):
            return (t[0] + (24 if t[0] < 4 else 0), t[1])
        ts.sort(key=key)
        first, last = ts[0], ts[-1]
        print(f"{st:<12}{d:<10}{len(ts):>5}  {first[0]:02d}:{first[1]:02d} {last[0]:02d}:{last[1]:02d}")


def _cell_hm(v):
    """xlsx 시각 셀 → (h, m). '05:02:00' 문자열 또는 datetime.time/datetime 모두 처리."""
    if v is None:
        return None
    if hasattr(v, "hour") and hasattr(v, "minute"):   # datetime.time / datetime
        return v.hour, v.minute
    m = re.match(r"^\s*(\d{1,2}):(\d{2})", str(v))
    return (int(m.group(1)), int(m.group(2))) if m else None


def parse_xlsx(path, line):
    """운영사 xlsx 시각표(코레일 계열) → 정규화 레코드. **line 을 넘겨줘야 한다**(파일이
    노선을 안 담음). 역명은 축약/변형이 심해(시흥청≠시흥시청·신김포 등) 그대로 두고,
    judge_subway 의 역명 재조정(reconcile)이 실측 역명에 맞춘다.

    포맷 (✅ 실측 4개 파일): matrix — 열=열차(열차번호), 행=역, 셀=시각(HH:MM:SS).
      시트명 = 방향(상행/하행) + 요일(평일/휴일). r0 시발역·r1 종착역·r2 열차번호(헤더).
      그 뒤 각 역이 **두 행**: 역명 행=도착, 다음 무명 행=출발. 도착(역명 행)을 쓴다
      (기점은 도착이 비어 다음 출발 행으로 대체). '연계열번' 등 라벨 행은 건너뛴다.
    """
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    out = []
    SKIP = {"시발역", "종착역", "열차번호", "연계열번", ""}
    for sh in wb.sheetnames:
        dtn = "공휴일" if ("휴일" in sh) else "평일"
        direction = "상행" if "상행" in sh else "하행" if "하행" in sh else None
        if direction is None:
            continue                              # 광운대시종착 등 방향 불명 시트는 제외
        rows = list(wb[sh].iter_rows(values_only=True))
        # 헤더(열차번호 행) 위치 — 그 아래부터 역 행
        hdr = next((i for i, r in enumerate(rows)
                    if r and str(r[0]).strip() == "열차번호"), 2)
        i = hdr + 1
        while i < len(rows):
            name = str(rows[i][0]).strip() if rows[i] and rows[i][0] is not None else ""
            if name in SKIP:
                i += 1
                continue
            times = [_cell_hm(c) for c in rows[i][1:]]
            # 도착 행이 전부 비면(기점) 다음 출발 행 사용
            if not any(times) and i + 1 < len(rows):
                times = [_cell_hm(c) for c in rows[i + 1][1:]]
            for hm in times:
                if hm:
                    out.append({"line": line, "station": name, "daytype": dtn,
                                "direction": direction, "updn": None,
                                "h": hm[0], "m": hm[1], "dest": None})
            # 역은 두 행(도착/출발)이므로 다음 무명 행을 건너뛴다
            j = i + 1
            if j < len(rows):
                nxt = str(rows[j][0]).strip() if rows[j] and rows[j][0] is not None else ""
                if nxt == "":
                    i = j + 1
                    continue
            i += 1
    wb.close()
    return out


def parse_stcs(csv_path):
    """서울교통공사 15098251 CSV → 정규화 레코드 (1~9호선).

    ⚠️ CSV 실물이 있어야 검증된다 (현재 저장소에 없음 — docstring 상단 참조).
    문서화된 스키마(§3.3.1, cp949):
      고유번호,호선,역사코드,역사명,주중주말,방향,급행여부,열차코드,
      열차도착시간,열차출발시간,출발역,도착역
      주중주말: DAY(평일)/SAT(토)/END(휴일)
    """
    import csv
    daymap = {"DAY": "평일", "SAT": "토요일", "END": "공휴일"}
    out = []
    with open(csv_path, encoding="cp949", newline="") as f:
        reader = csv.DictReader(f)
        # 헤더 명이 파일마다 미세하게 다를 수 있어(공백·괄호) 유연 매칭한다
        cols = {c: c for c in (reader.fieldnames or [])}
        def col(*names):
            for n in names:
                for c in cols:
                    if n in c.replace(" ", ""):
                        return c
            return None
        c_line = col("호선"); c_stat = col("역사명", "역명")
        c_day = col("주중주말", "요일"); c_dir = col("방향", "상하")
        c_arr = col("도착시간", "도착시각"); c_dep = col("출발시간", "출발시각")
        c_code = col("열차코드", "열차번호")
        if not all((c_line, c_stat, c_day, c_arr or c_dep)):
            sys.exit(f"15098251 컬럼을 못 찾았다 — 실제 헤더: {reader.fieldnames}\n"
                     "docstring 의 스키마와 대조해 col() 매칭을 조정할 것.")
        for row in reader:
            # 정시성은 실측 도착(arvlCd=1)과 대조하므로 **도착시간 우선**, 없으면 출발
            # (기점은 도착이 비고 출발만 있다). judge_subway 가 이 시각을 계획으로 쓴다.
            t = (row.get(c_arr) or row.get(c_dep) or "").strip()
            m = re.match(r"^(\d{1,2}):(\d{2})", t)
            if not m:
                continue
            out.append({
                "line": (row.get(c_line) or "").strip(),
                "station": (row.get(c_stat) or "").strip(),
                "daytype": daymap.get((row.get(c_day) or "").strip(),
                                      (row.get(c_day) or "").strip()),
                "direction": (row.get(c_dir) or "").strip() if c_dir else None,
                "updn": None,
                "h": int(m.group(1)), "m": int(m.group(2)),
                "dest": None,
                "traincode": (row.get(c_code) or "").strip() if c_code else None,
            })
    return out


def main():
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    cmd = sys.argv[1]
    args = [a for a in sys.argv[2:] if not a.startswith("--")]
    out_json = None
    if "--json" in sys.argv:
        out_json = sys.argv[sys.argv.index("--json") + 1]

    if cmd == "sinbundang":
        # daytype 은 파일마다 지정. 여러 파일이면 파일명으로 평일/공휴일 자동 추정.
        records = []
        forced = (sys.argv[sys.argv.index("--daytype") + 1]
                  if "--daytype" in sys.argv else None)
        pdfs = [a for a in args if a.endswith(".pdf")]
        for p in pdfs:
            dt = forced or ("공휴일" if "공휴일" in p else "토요일" if "토요일" in p else "평일")
            records += parse_sinbundang(p, dt)
        summarize(records)
    elif cmd == "stcs":
        records = parse_stcs(args[0])
        summarize(records)
    else:
        sys.exit(f"모르는 명령: {cmd}  (sinbundang | stcs)")

    if out_json:
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(records, f, ensure_ascii=False)
        print(f"\n→ {out_json} ({len(records):,}건)")


if __name__ == "__main__":
    main()
